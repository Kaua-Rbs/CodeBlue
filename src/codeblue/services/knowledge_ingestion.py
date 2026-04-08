from __future__ import annotations

import csv
import re
import unicodedata
from collections.abc import Iterable
from pathlib import Path
from typing import Any, Protocol

from codeblue.domain.knowledge_ingestion_models import (
    ActionLibraryRow,
    DeploymentSeasonalityProfileRow,
    InfluenzaPackInterventionRow,
    InfluenzaPackRiskFeatureRow,
    InfluenzaPackTimingRow,
    KnowledgeAbbreviationEntry,
    KnowledgeCsvTableSummary,
    KnowledgeEvidenceRow,
    KnowledgeLibraryEntry,
    KnowledgeSheetSummary,
    KnowledgeSourceCsvPackage,
    KnowledgeSourceImport,
    KnowledgeSourceRole,
    KnowledgeSourceSchemaFamily,
    KnowledgeSynthesisRow,
    KnowledgeTriggerEntry,
    PolicySourceRow,
    PolicyTriggerRow,
    TriggerActionMapRow,
)

EVIDENCE_CORE_HEADERS = {
    "source_id",
    "citation",
    "pathogen_scope",
    "canonical_feature_name",
    "codeblue_translation",
}
LIBRARY_HEADERS = {
    "canonical_feature_name",
    "alias_name",
    "feature_class",
    "default_temporal_stage",
}
ABBREVIATION_HEADERS = {"abbreviation", "meaning"}

COMPACT_HEADER_COUNT = 33

HEADER_ALIASES = {
    "pmid_doi": "pmid_doi",
    "pmid_doi_": "pmid_doi",
    "country_setting_population_study_design": "country_setting_population_study_design",
}

EVIDENCE_ROW_FIELDS = {
    "source_id",
    "citation",
    "pmid",
    "doi",
    "country",
    "setting",
    "population",
    "study_design",
    "pathogen_scope",
    "acquisition_context",
    "hospital_acquired_definition",
    "diagnostic_method",
    "surveillance_frame",
    "comparator_group",
    "canonical_feature_name",
    "risk_factor_type",
    "feature_class",
    "temporal_stage",
    "risk_factor_definition",
    "finding_text",
    "effect_size",
    "effect_size_type",
    "confidence_interval",
    "p_value",
    "effect_direction",
    "evidence_strength",
    "adjustment_level",
    "outcome_linked",
    "clinical_context",
    "applicability_to_influenza",
    "key_limitations",
    "study_reported_implication",
    "codeblue_translation",
    "tags",
    "factor_role",
    "data_status",
    "risk_factor_name",
}

CSV_STRUCTURED_TABLE_SPECS: dict[
    str,
    tuple[set[str], type[Any], str, str],
] = {
    "influenza_pack_timing": (
        {
            "pack_id",
            "timing_parameter",
            "parameter_group",
            "default_value",
            "data_status",
        },
        InfluenzaPackTimingRow,
        "influenza_pack_timing_rows",
        "influenza_pack_timing",
    ),
    "influenza_pack_risk_features": (
        {
            "pack_id",
            "canonical_feature_name",
            "priority_tier",
            "default_use_stage",
            "data_status",
        },
        InfluenzaPackRiskFeatureRow,
        "influenza_pack_risk_feature_rows",
        "influenza_pack_risk_features",
    ),
    "influenza_pack_interventions": (
        {
            "pack_id",
            "intervention_name",
            "action_type",
            "default_review_level",
            "data_status",
        },
        InfluenzaPackInterventionRow,
        "influenza_pack_intervention_rows",
        "influenza_pack_interventions",
    ),
    "deployment_seasonality_profile": (
        {
            "seasonality_profile_id",
            "hospital_id",
            "geography_label",
            "profile_status",
            "data_status",
        },
        DeploymentSeasonalityProfileRow,
        "deployment_profiles",
        "deployment_seasonality_profile",
    ),
    "policy_source_library": (
        {
            "policy_source_id",
            "source_title",
            "issuing_body",
            "source_type",
            "data_status",
        },
        PolicySourceRow,
        "policy_sources",
        "policy_source_library",
    ),
    "policy_trigger_library": (
        {
            "trigger_id",
            "trigger_name",
            "trigger_type",
            "logic_definition",
            "status",
        },
        PolicyTriggerRow,
        "policy_triggers",
        "policy_trigger_library",
    ),
    "action_library": (
        {
            "action_id",
            "action_name",
            "action_domain",
            "action_intent",
            "data_status",
        },
        ActionLibraryRow,
        "action_library_rows",
        "action_library",
    ),
    "trigger_action_map": (
        {
            "map_id",
            "trigger_id",
            "action_id",
            "relationship_type",
            "data_status",
        },
        TriggerActionMapRow,
        "trigger_action_map_rows",
        "trigger_action_map",
    ),
}


class Worksheet(Protocol):
    title: str

    def iter_rows(
        self,
        min_row: int = 1,
        values_only: bool = False,
    ) -> Iterable[tuple[Any, ...]]: ...


class Workbook(Protocol):
    sheetnames: list[str]

    def __getitem__(self, key: str) -> Worksheet: ...


def load_knowledge_source_workbook(path: Path) -> KnowledgeSourceImport:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    workbook = load_workbook(filename=path, data_only=True)
    return parse_knowledge_source_workbook(workbook, source_name=path.name)


def load_knowledge_source_csv_package(path: Path) -> KnowledgeSourceCsvPackage:
    if not path.exists():
        raise ValueError(f"Knowledge source directory {path!s} does not exist.")
    if not path.is_dir():
        raise ValueError(f"Knowledge source path {path!s} is not a directory.")

    imported = KnowledgeSourceCsvPackage(source_directory=str(path))

    for csv_path in sorted(path.glob("*.csv")):
        table_name = normalize_source_package_table_name(csv_path.stem)

        if table_name in CSV_STRUCTURED_TABLE_SPECS:
            summary, structured_rows, target_attr = parse_structured_source_csv(
                csv_path, table_name
            )
            imported.tables.append(summary)
            getattr(imported, target_attr).extend(structured_rows)
            continue

        if table_name == "library":
            summary, library_entries = parse_knowledge_library_csv(csv_path)
            imported.tables.append(summary)
            imported.library_entries.extend(library_entries)
            continue

        if table_name == "abreviacoes":
            summary, abbreviation_entries = parse_knowledge_abbreviation_csv(csv_path)
            imported.tables.append(summary)
            imported.abbreviation_entries.extend(abbreviation_entries)
            continue

        try:
            summary, evidence_rows = parse_knowledge_source_evidence_csv(csv_path)
        except ValueError:
            imported.unmodeled_files.append(csv_path.name)
            continue

        imported.tables.append(summary)
        imported.evidence_rows.extend(evidence_rows)

    return imported


def parse_knowledge_source_workbook(
    workbook: Workbook, source_name: str = "HSIL.xlsx"
) -> KnowledgeSourceImport:
    imported = KnowledgeSourceImport(source_name=source_name)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        role = classify_knowledge_source_sheet(sheet_name)
        schema_family = classify_knowledge_source_schema_family(sheet_name)

        if role == KnowledgeSourceRole.PLACEHOLDER:
            imported.sheets.append(
                KnowledgeSheetSummary(
                    sheet_name=sheet_name,
                    role=role,
                    schema_family=schema_family,
                )
            )
            continue

        if role == KnowledgeSourceRole.OVERVIEW_SYNTHESIS:
            summary, synthesis_rows = parse_knowledge_source_synthesis_sheet(worksheet)
            imported.sheets.append(summary)
            imported.synthesis_rows.extend(synthesis_rows)
            continue

        if (
            role == KnowledgeSourceRole.REFERENCE
            and normalize_sheet_name(sheet_name) == "abreviacoes"
        ):
            summary, abbreviation_entries = parse_knowledge_abbreviation_sheet(worksheet)
            imported.sheets.append(summary)
            imported.abbreviation_entries.extend(abbreviation_entries)
            continue

        if role == KnowledgeSourceRole.REFERENCE and normalize_sheet_name(sheet_name) == "library":
            summary, library_entries = parse_knowledge_library_sheet(worksheet)
            imported.sheets.append(summary)
            imported.library_entries.extend(library_entries)
            continue

        if role == KnowledgeSourceRole.REFERENCE and "triggers" in normalize_sheet_name(sheet_name):
            summary, trigger_entries = parse_knowledge_trigger_sheet(worksheet)
            imported.sheets.append(summary)
            imported.trigger_entries.extend(trigger_entries)
            continue

        summary, evidence_rows = parse_knowledge_source_evidence_sheet(worksheet, role)
        imported.sheets.append(summary)
        imported.evidence_rows.extend(evidence_rows)

    return imported


def parse_structured_source_csv(
    path: Path,
    table_name: str,
) -> tuple[KnowledgeCsvTableSummary, list[Any], str]:
    required_headers, row_model, target_attr, handled_as = CSV_STRUCTURED_TABLE_SPECS[table_name]
    rows = read_csv_rows(path)
    header_row_index, headers = locate_csv_header_row(rows, required_headers=required_headers)
    row_dicts, duplicate_headers = extract_csv_row_dicts(rows, header_row_index, headers)
    parsed_rows = [row_model.model_validate(row) for row in row_dicts]
    return (
        KnowledgeCsvTableSummary(
            file_name=path.name,
            table_name=table_name,
            handled_as=handled_as,
            row_count=len(parsed_rows),
            header_row_index=header_row_index + 1,
            skipped_title_rows=header_row_index,
            skipped_duplicate_header_rows=duplicate_headers,
        ),
        parsed_rows,
        target_attr,
    )


def parse_knowledge_source_evidence_csv(
    path: Path,
) -> tuple[KnowledgeCsvTableSummary, list[KnowledgeEvidenceRow]]:
    rows = read_csv_rows(path)
    header_row_index, headers = locate_csv_header_row(rows)
    schema_family = infer_knowledge_table_schema_family(headers)
    row_dicts, duplicate_headers = extract_csv_row_dicts(rows, header_row_index, headers)
    parsed_rows = [
        KnowledgeEvidenceRow(
            sheet_name=path.stem,
            schema_family=schema_family,
            **{key: value for key, value in row.items() if key in EVIDENCE_ROW_FIELDS},
        )
        for row in row_dicts
    ]
    return (
        KnowledgeCsvTableSummary(
            file_name=path.name,
            table_name=normalize_source_package_table_name(path.stem),
            handled_as="evidence_rows",
            row_count=len(parsed_rows),
            header_row_index=header_row_index + 1,
            skipped_title_rows=header_row_index,
            skipped_duplicate_header_rows=duplicate_headers,
        ),
        parsed_rows,
    )


def parse_knowledge_library_csv(
    path: Path,
) -> tuple[KnowledgeCsvTableSummary, list[KnowledgeLibraryEntry]]:
    rows = read_csv_rows(path)
    header_row_index, headers = locate_csv_header_row(rows, required_headers=LIBRARY_HEADERS)
    row_dicts, duplicate_headers = extract_csv_row_dicts(rows, header_row_index, headers)
    entries: list[KnowledgeLibraryEntry] = []
    for row in row_dicts:
        canonical_feature_name = row.get("canonical_feature_name")
        if not canonical_feature_name:
            continue
        entries.append(
            KnowledgeLibraryEntry(
                canonical_feature_name=canonical_feature_name,
                alias_name=row.get("alias_name"),
                feature_class=row.get("feature_class"),
                default_temporal_stage=row.get("default_temporal_stage"),
                description=row.get("description"),
                typical_effect_direction=row.get("typical_effect_direction"),
                deployment_priority=row.get("deployment_priority"),
            )
        )
    return (
        KnowledgeCsvTableSummary(
            file_name=path.name,
            table_name="library",
            handled_as="library_entries",
            row_count=len(entries),
            header_row_index=header_row_index + 1,
            skipped_title_rows=header_row_index,
            skipped_duplicate_header_rows=duplicate_headers,
        ),
        entries,
    )


def parse_knowledge_abbreviation_csv(
    path: Path,
) -> tuple[KnowledgeCsvTableSummary, list[KnowledgeAbbreviationEntry]]:
    rows = read_csv_rows(path)
    entries: list[KnowledgeAbbreviationEntry] = []
    for row in rows[1:]:
        if not row:
            continue
        cell = normalize_cell(row[0]) if row[0:] else ""
        if not cell:
            continue
        parts = re.split(r"\s+[—-]\s+", cell, maxsplit=1)
        if len(parts) != 2:
            continue
        entries.append(
            KnowledgeAbbreviationEntry(
                abbreviation=parts[0].strip(),
                meaning=parts[1].strip(),
            )
        )
    return (
        KnowledgeCsvTableSummary(
            file_name=path.name,
            table_name="abreviacoes",
            handled_as="abbreviation_entries",
            row_count=len(entries),
            header_row_index=1,
            skipped_title_rows=0,
            skipped_duplicate_header_rows=0,
        ),
        entries,
    )


def parse_knowledge_source_synthesis_sheet(
    worksheet: Worksheet,
) -> tuple[KnowledgeSheetSummary, list[KnowledgeSynthesisRow]]:
    header_row_index, headers = locate_generic_header_row(worksheet)
    row_dicts, duplicate_headers = extract_row_dicts(worksheet, header_row_index, headers)
    rows = [
        KnowledgeSynthesisRow(
            sheet_name=worksheet.title,
            row_payload={key: value for key, value in row.items() if key and value},
        )
        for row in row_dicts
    ]
    summary = KnowledgeSheetSummary(
        sheet_name=worksheet.title,
        role=KnowledgeSourceRole.OVERVIEW_SYNTHESIS,
        schema_family=KnowledgeSourceSchemaFamily.SYNTHESIS,
        header_row_index=header_row_index + 1,
        skipped_title_rows=header_row_index,
        skipped_duplicate_header_rows=duplicate_headers,
        data_row_count=len(rows),
    )
    return summary, rows


def parse_knowledge_source_evidence_sheet(
    worksheet: Worksheet, role: KnowledgeSourceRole
) -> tuple[KnowledgeSheetSummary, list[KnowledgeEvidenceRow]]:
    header_row_index, headers = locate_header_row(worksheet)
    schema_family = infer_knowledge_table_schema_family(headers)
    row_dicts, duplicate_headers = extract_row_dicts(worksheet, header_row_index, headers)

    parsed_rows = [
        KnowledgeEvidenceRow(
            sheet_name=worksheet.title,
            schema_family=schema_family,
            **{key: value for key, value in row.items() if key in EVIDENCE_ROW_FIELDS},
        )
        for row in row_dicts
    ]

    summary = KnowledgeSheetSummary(
        sheet_name=worksheet.title,
        role=role,
        schema_family=schema_family,
        header_row_index=header_row_index + 1,
        skipped_title_rows=header_row_index,
        skipped_duplicate_header_rows=duplicate_headers,
        data_row_count=len(parsed_rows),
    )
    return summary, parsed_rows


def parse_knowledge_library_sheet(
    worksheet: Worksheet,
) -> tuple[KnowledgeSheetSummary, list[KnowledgeLibraryEntry]]:
    header_row_index, headers = locate_header_row(worksheet, required_headers=LIBRARY_HEADERS)
    row_dicts, duplicate_headers = extract_row_dicts(worksheet, header_row_index, headers)
    entries: list[KnowledgeLibraryEntry] = []
    for row in row_dicts:
        canonical_feature_name = row.get("canonical_feature_name")
        if not canonical_feature_name:
            continue
        entries.append(
            KnowledgeLibraryEntry(
                canonical_feature_name=canonical_feature_name,
                alias_name=row.get("alias_name"),
                feature_class=row.get("feature_class"),
                default_temporal_stage=row.get("default_temporal_stage"),
                description=row.get("description"),
                typical_effect_direction=row.get("typical_effect_direction"),
                deployment_priority=row.get("deployment_priority"),
            )
        )
    summary = KnowledgeSheetSummary(
        sheet_name=worksheet.title,
        role=KnowledgeSourceRole.REFERENCE,
        schema_family=KnowledgeSourceSchemaFamily.REFERENCE,
        header_row_index=header_row_index + 1,
        skipped_title_rows=header_row_index,
        skipped_duplicate_header_rows=duplicate_headers,
        data_row_count=len(entries),
    )
    return summary, entries


def parse_knowledge_trigger_sheet(
    worksheet: Worksheet,
) -> tuple[KnowledgeSheetSummary, list[KnowledgeTriggerEntry]]:
    values = iter_sheet_values(worksheet)
    entries = [KnowledgeTriggerEntry(trigger_text=value) for value in values if value]
    summary = KnowledgeSheetSummary(
        sheet_name=worksheet.title,
        role=KnowledgeSourceRole.REFERENCE,
        schema_family=KnowledgeSourceSchemaFamily.REFERENCE,
        data_row_count=len(entries),
    )
    return summary, entries


def parse_knowledge_abbreviation_sheet(
    worksheet: Worksheet,
) -> tuple[KnowledgeSheetSummary, list[KnowledgeAbbreviationEntry]]:
    header_row_index, headers = locate_header_row(worksheet, required_headers=ABBREVIATION_HEADERS)
    row_dicts, duplicate_headers = extract_row_dicts(worksheet, header_row_index, headers)
    entries = [
        KnowledgeAbbreviationEntry(
            abbreviation=row["abbreviation"],
            meaning=row["meaning"],
        )
        for row in row_dicts
        if row.get("abbreviation") and row.get("meaning")
    ]
    summary = KnowledgeSheetSummary(
        sheet_name=worksheet.title,
        role=KnowledgeSourceRole.REFERENCE,
        schema_family=KnowledgeSourceSchemaFamily.REFERENCE,
        header_row_index=header_row_index + 1,
        skipped_title_rows=header_row_index,
        skipped_duplicate_header_rows=duplicate_headers,
        data_row_count=len(entries),
    )
    return summary, entries


def classify_knowledge_source_sheet(sheet_name: str) -> KnowledgeSourceRole:
    normalized = normalize_sheet_name(sheet_name)
    if normalized == "riscos_2":
        return KnowledgeSourceRole.OVERVIEW_SYNTHESIS
    if normalized in {
        "ha_influenza_risk_factors",
        "hospitalized_influenza_severity",
        "secondary_infections_in_influen",
        "pathogen_interactions",
    }:
        return KnowledgeSourceRole.CORE_EVIDENCE
    if normalized in {
        "transmission_context_support",
        "staff_vector_transmission",
        "advanced_outbreak_monitoring_to",
    }:
        return KnowledgeSourceRole.COMPANION_EVIDENCE
    if normalized in {"library", "survaillence_info_triggers", "abreviacoes"}:
        return KnowledgeSourceRole.REFERENCE
    if normalized == "pathogen_info":
        return KnowledgeSourceRole.PLACEHOLDER
    return KnowledgeSourceRole.REFERENCE


def classify_knowledge_source_schema_family(sheet_name: str) -> KnowledgeSourceSchemaFamily:
    normalized = normalize_sheet_name(sheet_name)
    if normalized == "riscos_2":
        return KnowledgeSourceSchemaFamily.SYNTHESIS
    if normalized == "ha_influenza_risk_factors":
        return KnowledgeSourceSchemaFamily.COMPACT_CORE
    if normalized in {"transmission_context_support", "staff_vector_transmission"}:
        return KnowledgeSourceSchemaFamily.EXTENDED
    if normalized in {
        "hospitalized_influenza_severity",
        "secondary_infections_in_influen",
        "pathogen_interactions",
        "advanced_outbreak_monitoring_to",
    }:
        return KnowledgeSourceSchemaFamily.LEGACY_EXTENDED
    if normalized == "pathogen_info":
        return KnowledgeSourceSchemaFamily.PLACEHOLDER
    return KnowledgeSourceSchemaFamily.REFERENCE


def locate_header_row(
    worksheet: Worksheet, required_headers: set[str] | None = None
) -> tuple[int, list[str]]:
    expected = required_headers or EVIDENCE_CORE_HEADERS
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        headers = [normalize_knowledge_header(value) for value in row]
        if expected.issubset(set(headers)):
            return index, headers
    raise ValueError(f"Could not locate header row for worksheet {worksheet.title!r}.")


def locate_csv_header_row(
    rows: list[list[str]],
    required_headers: set[str] | None = None,
) -> tuple[int, list[str]]:
    expected = required_headers or EVIDENCE_CORE_HEADERS
    for index, row in enumerate(rows):
        headers = [normalize_knowledge_header(value) for value in row]
        if expected.issubset(set(headers)):
            return index, headers
    raise ValueError("Could not locate a matching CSV header row.")


def locate_generic_header_row(worksheet: Worksheet) -> tuple[int, list[str]]:
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        headers = [normalize_knowledge_header(value) for value in row]
        if len([header for header in headers if header]) >= 4:
            return index, headers
    raise ValueError(f"Could not locate generic header row for worksheet {worksheet.title!r}.")


def extract_row_dicts(
    worksheet: Worksheet, header_row_index: int, headers: list[str]
) -> tuple[list[dict[str, str]], int]:
    rows: list[dict[str, str]] = []
    duplicate_headers = 0

    for row in worksheet.iter_rows(min_row=header_row_index + 2, values_only=True):
        values = [normalize_cell(value) for value in row[: len(headers)]]
        if not any(values):
            continue
        if [normalize_knowledge_header(value) for value in row[: len(headers)]] == headers:
            duplicate_headers += 1
            continue
        mapped = {header: values[idx] for idx, header in enumerate(headers) if header}
        rows.append(mapped)
    return rows, duplicate_headers


def extract_csv_row_dicts(
    rows: list[list[str]],
    header_row_index: int,
    headers: list[str],
) -> tuple[list[dict[str, str | None]], int]:
    extracted: list[dict[str, str | None]] = []
    duplicate_headers = 0

    for row in rows[header_row_index + 1 :]:
        values = [
            normalize_source_cell(row[idx] if idx < len(row) else None)
            for idx in range(len(headers))
        ]
        if not any(value is not None for value in values):
            continue
        normalized_row_headers = [
            normalize_knowledge_header(row[idx] if idx < len(row) else None)
            for idx in range(len(headers))
        ]
        if normalized_row_headers == headers:
            duplicate_headers += 1
            continue
        mapped = {header: values[idx] for idx, header in enumerate(headers) if header}
        extracted.append(mapped)
    return extracted, duplicate_headers


def infer_knowledge_table_schema_family(headers: list[str]) -> KnowledgeSourceSchemaFamily:
    header_set = set(headers)
    if "risk_factor_name" in header_set:
        return KnowledgeSourceSchemaFamily.LEGACY_EXTENDED
    if {"tags", "factor_role", "data_status"}.issubset(header_set):
        return KnowledgeSourceSchemaFamily.EXTENDED
    if len([header for header in headers if header]) >= COMPACT_HEADER_COUNT:
        return KnowledgeSourceSchemaFamily.COMPACT_CORE
    return KnowledgeSourceSchemaFamily.REFERENCE


def iter_sheet_values(worksheet: Worksheet) -> list[str]:
    values: list[str] = []
    for row in worksheet.iter_rows(values_only=True):
        for cell in row:
            normalized = normalize_cell(cell)
            if normalized:
                values.append(normalized)
    return values


def normalize_sheet_name(name: str) -> str:
    return (
        normalize_knowledge_header(name)
        .replace("transmission_context_suppo_rt", "transmission_context_support")
        .replace("transmission_context_suppo", "transmission_context_support")
        .replace("abreviacoes", "abreviacoes")
    )


def normalize_source_package_table_name(name: str) -> str:
    normalized = normalize_sheet_name(name)
    if normalized.startswith("hsil_"):
        normalized = normalized.removeprefix("hsil_")
    return normalized


def normalize_knowledge_header(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    text = text.strip().lower()
    text = re.sub(r"[\/\-\.,:()]+", "_", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return HEADER_ALIASES.get(text, text)


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_source_cell(value: Any) -> str | None:
    normalized = normalize_cell(value)
    if not normalized:
        return None
    if normalized.lower() == "blank":
        return None
    return normalized


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.reader(handle))
